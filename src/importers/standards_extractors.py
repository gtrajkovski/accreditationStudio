"""Standards extraction from multiple file formats.

Factory pattern for pluggable extractors. Each extractor takes a file
and returns raw text with structural hints (headers, tables, etc.).
"""

from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Dict, Any, List, Optional
import logging
import re
import csv
import io

logger = logging.getLogger(__name__)


class ExtractorType(str, Enum):
    """Supported extractor types."""
    PDF = "pdf"
    EXCEL = "excel"
    CSV = "csv"
    TEXT = "text"
    WEB = "web"


@dataclass
class ExtractedContent:
    """Result of content extraction."""
    source_type: ExtractorType
    source_path: str
    raw_text: str
    structural_hints: List[Dict[str, Any]] = field(default_factory=list)
    # structural_hints: [{"type": "heading", "level": 1, "text": "...", "position": n}, ...]
    tables: List[Dict[str, Any]] = field(default_factory=list)
    # tables: [{"headers": [...], "rows": [[...]], "position": n}, ...]
    metadata: Dict[str, Any] = field(default_factory=dict)
    page_count: int = 1
    errors: List[str] = field(default_factory=list)
    confidence: float = 0.0  # 0-1 extraction confidence

    def to_dict(self) -> Dict[str, Any]:
        return {
            "source_type": self.source_type.value,
            "source_path": self.source_path,
            "raw_text": self.raw_text,
            "structural_hints": self.structural_hints,
            "tables": self.tables,
            "metadata": self.metadata,
            "page_count": self.page_count,
            "errors": self.errors,
            "confidence": self.confidence,
        }

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> "ExtractedContent":
        source_type = data.get("source_type", "text")
        if isinstance(source_type, str):
            try:
                source_type = ExtractorType(source_type)
            except ValueError:
                source_type = ExtractorType.TEXT

        return cls(
            source_type=source_type,
            source_path=data.get("source_path", ""),
            raw_text=data.get("raw_text", ""),
            structural_hints=data.get("structural_hints", []),
            tables=data.get("tables", []),
            metadata=data.get("metadata", {}),
            page_count=data.get("page_count", 1),
            errors=data.get("errors", []),
            confidence=data.get("confidence", 0.0),
        )


class BaseExtractor(ABC):
    """Abstract base class for content extractors."""

    # Common heading patterns
    HEADING_PATTERNS = [
        # Roman numerals: I, II, III, IV, V, VI, VII, VIII, IX, X, XI, XII
        (r"^(I{1,3}|IV|VI{0,3}|IX|X{1,3}|XI{1,3}|XII)\.\s*(.+)$", 1),
        # Combined Roman + Letter: I.A, II.B
        (r"^(I{1,3}|IV|VI{0,3}|IX|X{1,3})\.[A-Z]\.\s*(.+)$", 2),
        # Combined Roman + Letter + Number: I.A.1
        (r"^(I{1,3}|IV|VI{0,3}|IX|X{1,3})\.[A-Z]\.\d+\.\s*(.+)$", 3),
        # Arabic decimals: 1., 2., 3.
        (r"^(\d+)\.\s+(.+)$", 1),
        # Combined decimals: 1.1, 1.2, 2.1
        (r"^(\d+)\.(\d+)\s+(.+)$", 2),
        # Triple decimals: 1.1.1, 1.2.3
        (r"^(\d+)\.(\d+)\.(\d+)\s+(.+)$", 3),
        # Letter headings: A., B., C.
        (r"^([A-Z])\.\s+(.+)$", 2),
        # Parenthetical: (1), (2), (a), (b)
        (r"^\((\d+|[a-z]|[A-Z])\)\s+(.+)$", 3),
    ]

    @property
    @abstractmethod
    def extractor_type(self) -> ExtractorType:
        """Return the extractor type."""
        pass

    @abstractmethod
    def extract(self, source: str, **kwargs) -> ExtractedContent:
        """Extract content from source.

        Args:
            source: File path or URL
            **kwargs: Extractor-specific options

        Returns:
            ExtractedContent with raw text and structural hints
        """
        pass

    def _detect_headings(self, text: str) -> List[Dict[str, Any]]:
        """Detect heading patterns in text.

        Looks for common patterns:
        - Roman numerals: I, II, III
        - Lettered: A, B, C or (A), (B)
        - Numbered: 1, 2, 3 or 1., 2.
        - Combined: I.A.1, 1.2.3
        """
        headings = []
        lines = text.split('\n')

        for i, line in enumerate(lines):
            line = line.strip()
            if not line:
                continue

            for pattern, level in self.HEADING_PATTERNS:
                match = re.match(pattern, line, re.MULTILINE)
                if match:
                    headings.append({
                        "type": "heading",
                        "level": level,
                        "text": line,
                        "position": i,
                    })
                    break

        return headings

    def _calculate_confidence(self, text: str, headings: List[Dict], tables: List[Dict]) -> float:
        """Calculate extraction confidence based on content quality."""
        confidence = 0.5  # Base confidence

        # Boost for meaningful text length
        if len(text) > 1000:
            confidence += 0.1
        if len(text) > 5000:
            confidence += 0.1

        # Boost for detected structure
        if headings:
            confidence += 0.1
        if len(headings) > 5:
            confidence += 0.1

        # Boost for tables
        if tables:
            confidence += 0.1

        # Penalty for very short text
        if len(text) < 100:
            confidence -= 0.3

        return min(1.0, max(0.0, confidence))


class PdfExtractor(BaseExtractor):
    """Extract content from PDF files using pdfplumber."""

    @property
    def extractor_type(self) -> ExtractorType:
        return ExtractorType.PDF

    def extract(self, source: str, **kwargs) -> ExtractedContent:
        """Extract text and structure from PDF.

        Detects:
        - Page boundaries
        - Bold/large text (potential headings)
        - Tables
        - Lists
        """
        import pdfplumber

        path = Path(source)
        if not path.exists():
            return ExtractedContent(
                source_type=self.extractor_type,
                source_path=source,
                raw_text="",
                errors=[f"File not found: {source}"],
                confidence=0.0,
            )

        logger.info(f"Extracting content from PDF: {source}")

        pages_text = []
        all_tables = []
        page_count = 0
        errors = []

        try:
            with pdfplumber.open(path) as pdf:
                page_count = len(pdf.pages)

                for page_num, page in enumerate(pdf.pages):
                    # Extract text
                    try:
                        text = page.extract_text() or ""
                        pages_text.append(f"[Page {page_num + 1}]\n{text}")
                    except Exception as e:
                        logger.warning(f"Failed to extract text from page {page_num + 1}: {e}")
                        errors.append(f"Page {page_num + 1} text extraction failed: {e}")

                    # Extract tables
                    try:
                        tables = page.extract_tables()
                        for table in tables:
                            if table and len(table) > 1:
                                headers = table[0] if table else []
                                rows = table[1:] if len(table) > 1 else []
                                all_tables.append({
                                    "headers": headers,
                                    "rows": rows,
                                    "page": page_num + 1,
                                })
                    except Exception as e:
                        logger.warning(f"Failed to extract tables from page {page_num + 1}: {e}")

        except Exception as e:
            logger.error(f"Failed to parse PDF {source}: {e}")
            return ExtractedContent(
                source_type=self.extractor_type,
                source_path=source,
                raw_text="",
                errors=[f"PDF parse error: {e}"],
                confidence=0.0,
            )

        full_text = "\n\n".join(pages_text)

        # Detect headings
        headings = self._detect_headings(full_text)

        # Calculate confidence
        confidence = self._calculate_confidence(full_text, headings, all_tables)

        # Warn if text is sparse (possible image-based PDF)
        if len(full_text) < 100 and page_count > 0:
            errors.append("Very little text extracted. This may be an image-based PDF requiring OCR.")
            confidence *= 0.5

        logger.info(f"Extracted {len(full_text)} chars, {len(headings)} headings, {len(all_tables)} tables from {page_count} pages")

        return ExtractedContent(
            source_type=self.extractor_type,
            source_path=source,
            raw_text=full_text,
            structural_hints=headings,
            tables=all_tables,
            metadata={
                "file_name": path.name,
                "page_count": page_count,
                "extraction_method": "pdfplumber",
            },
            page_count=page_count,
            errors=errors,
            confidence=confidence,
        )


class ExcelExtractor(BaseExtractor):
    """Extract content from Excel files using openpyxl."""

    @property
    def extractor_type(self) -> ExtractorType:
        return ExtractorType.EXCEL

    def extract(self, source: str, sheet_name: Optional[str] = None, **kwargs) -> ExtractedContent:
        """Extract content from Excel.

        Assumes structure:
        - First row: headers (number, title, description, etc.)
        - Subsequent rows: standards/requirements
        - Multiple sheets = multiple sections
        """
        from openpyxl import load_workbook

        path = Path(source)
        if not path.exists():
            return ExtractedContent(
                source_type=self.extractor_type,
                source_path=source,
                raw_text="",
                errors=[f"File not found: {source}"],
                confidence=0.0,
            )

        logger.info(f"Extracting content from Excel: {source}")

        text_parts = []
        all_tables = []
        headings = []
        errors = []
        sheet_count = 0

        try:
            wb = load_workbook(path, data_only=True)
            sheets = [sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.sheetnames
            sheet_count = len(sheets)

            for sheet_idx, sheet in enumerate(sheets):
                ws = wb[sheet]

                # Add sheet as a heading
                headings.append({
                    "type": "heading",
                    "level": 1,
                    "text": f"Sheet: {sheet}",
                    "position": len(text_parts),
                })
                text_parts.append(f"\n[Sheet: {sheet}]\n")

                # Read all rows
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue

                # First row as headers
                headers = [str(cell) if cell else "" for cell in rows[0]]
                data_rows = []

                for row in rows[1:]:
                    row_data = [str(cell) if cell else "" for cell in row]
                    # Skip empty rows
                    if any(cell.strip() for cell in row_data):
                        data_rows.append(row_data)
                        # Build text representation
                        row_text = " | ".join(f"{h}: {v}" for h, v in zip(headers, row_data) if v.strip())
                        if row_text:
                            text_parts.append(row_text)

                if data_rows:
                    all_tables.append({
                        "headers": headers,
                        "rows": data_rows,
                        "sheet": sheet,
                    })

            wb.close()

        except Exception as e:
            logger.error(f"Failed to parse Excel {source}: {e}")
            return ExtractedContent(
                source_type=self.extractor_type,
                source_path=source,
                raw_text="",
                errors=[f"Excel parse error: {e}"],
                confidence=0.0,
            )

        full_text = "\n".join(text_parts)

        # Calculate confidence
        confidence = self._calculate_confidence(full_text, headings, all_tables)

        # Excel files with structure get higher confidence
        if all_tables:
            confidence = min(1.0, confidence + 0.2)

        logger.info(f"Extracted {len(full_text)} chars from {sheet_count} sheets, {len(all_tables)} tables")

        return ExtractedContent(
            source_type=self.extractor_type,
            source_path=source,
            raw_text=full_text,
            structural_hints=headings,
            tables=all_tables,
            metadata={
                "file_name": path.name,
                "sheet_count": sheet_count,
                "extraction_method": "openpyxl",
            },
            page_count=sheet_count,
            errors=errors,
            confidence=confidence,
        )


class CsvExtractor(BaseExtractor):
    """Extract content from CSV files."""

    @property
    def extractor_type(self) -> ExtractorType:
        return ExtractorType.CSV

    def extract(self, source: str, delimiter: str = ",", **kwargs) -> ExtractedContent:
        """Extract content from CSV.

        Similar to Excel but single table only.
        """
        path = Path(source)
        if not path.exists():
            return ExtractedContent(
                source_type=self.extractor_type,
                source_path=source,
                raw_text="",
                errors=[f"File not found: {source}"],
                confidence=0.0,
            )

        logger.info(f"Extracting content from CSV: {source}")

        text_parts = []
        tables = []
        errors = []

        try:
            # Detect encoding
            with open(path, 'rb') as f:
                raw_content = f.read()

            # Try UTF-8 first, fall back to latin-1
            try:
                content = raw_content.decode('utf-8')
            except UnicodeDecodeError:
                content = raw_content.decode('latin-1')

            # Parse CSV
            reader = csv.reader(io.StringIO(content), delimiter=delimiter)
            rows = list(reader)

            if not rows:
                return ExtractedContent(
                    source_type=self.extractor_type,
                    source_path=source,
                    raw_text="",
                    errors=["Empty CSV file"],
                    confidence=0.0,
                )

            # First row as headers
            headers = rows[0]
            data_rows = rows[1:]

            # Build text representation
            for row in data_rows:
                if any(cell.strip() for cell in row):
                    row_text = " | ".join(f"{h}: {v}" for h, v in zip(headers, row) if v.strip())
                    if row_text:
                        text_parts.append(row_text)

            tables.append({
                "headers": headers,
                "rows": data_rows,
            })

        except Exception as e:
            logger.error(f"Failed to parse CSV {source}: {e}")
            return ExtractedContent(
                source_type=self.extractor_type,
                source_path=source,
                raw_text="",
                errors=[f"CSV parse error: {e}"],
                confidence=0.0,
            )

        full_text = "\n".join(text_parts)

        # Calculate confidence
        confidence = self._calculate_confidence(full_text, [], tables)

        # CSV with good structure gets higher confidence
        if tables and len(data_rows) > 5:
            confidence = min(1.0, confidence + 0.2)

        logger.info(f"Extracted {len(full_text)} chars, {len(data_rows)} rows from CSV")

        return ExtractedContent(
            source_type=self.extractor_type,
            source_path=source,
            raw_text=full_text,
            structural_hints=[],
            tables=tables,
            metadata={
                "file_name": path.name,
                "row_count": len(data_rows),
                "column_count": len(headers),
                "extraction_method": "csv",
            },
            page_count=1,
            errors=errors,
            confidence=confidence,
        )


class TextExtractor(BaseExtractor):
    """Extract content from plain text files."""

    @property
    def extractor_type(self) -> ExtractorType:
        return ExtractorType.TEXT

    def extract(self, source: str, **kwargs) -> ExtractedContent:
        """Extract content from text file.

        Applies heading detection patterns.
        """
        path = Path(source)
        if not path.exists():
            return ExtractedContent(
                source_type=self.extractor_type,
                source_path=source,
                raw_text="",
                errors=[f"File not found: {source}"],
                confidence=0.0,
            )

        logger.info(f"Extracting content from text file: {source}")

        errors = []

        try:
            # Try UTF-8 first, fall back to latin-1
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    content = f.read()
            except UnicodeDecodeError:
                with open(path, 'r', encoding='latin-1') as f:
                    content = f.read()

        except Exception as e:
            logger.error(f"Failed to read text file {source}: {e}")
            return ExtractedContent(
                source_type=self.extractor_type,
                source_path=source,
                raw_text="",
                errors=[f"File read error: {e}"],
                confidence=0.0,
            )

        # Detect headings
        headings = self._detect_headings(content)

        # Calculate confidence
        confidence = self._calculate_confidence(content, headings, [])

        logger.info(f"Extracted {len(content)} chars, {len(headings)} headings from text file")

        return ExtractedContent(
            source_type=self.extractor_type,
            source_path=source,
            raw_text=content,
            structural_hints=headings,
            tables=[],
            metadata={
                "file_name": path.name,
                "extraction_method": "text",
            },
            page_count=1,
            errors=errors,
            confidence=confidence,
        )


class WebExtractor(BaseExtractor):
    """Extract content from web pages."""

    @property
    def extractor_type(self) -> ExtractorType:
        return ExtractorType.WEB

    def extract(self, source: str, **kwargs) -> ExtractedContent:
        """Extract content from URL.

        Uses requests to fetch and BeautifulSoup for parsing if available,
        otherwise falls back to basic HTML stripping.
        """
        import requests

        logger.info(f"Extracting content from URL: {source}")

        errors = []

        try:
            response = requests.get(source, timeout=30, headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AccreditAI Standards Importer"
            })
            response.raise_for_status()
            html_content = response.text

        except Exception as e:
            logger.error(f"Failed to fetch URL {source}: {e}")
            return ExtractedContent(
                source_type=self.extractor_type,
                source_path=source,
                raw_text="",
                errors=[f"URL fetch error: {e}"],
                confidence=0.0,
            )

        # Try BeautifulSoup if available
        try:
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(html_content, 'html.parser')

            # Remove script and style elements
            for element in soup(['script', 'style', 'nav', 'footer', 'header']):
                element.decompose()

            # Extract text
            content = soup.get_text(separator='\n', strip=True)

            # Extract headings from HTML structure
            headings = []
            for level in range(1, 7):
                for i, heading in enumerate(soup.find_all(f'h{level}')):
                    headings.append({
                        "type": "heading",
                        "level": level,
                        "text": heading.get_text(strip=True),
                        "position": i,
                    })

            # Extract tables
            tables = []
            for table in soup.find_all('table'):
                rows = table.find_all('tr')
                if rows:
                    headers = [th.get_text(strip=True) for th in rows[0].find_all(['th', 'td'])]
                    data_rows = []
                    for row in rows[1:]:
                        data_rows.append([td.get_text(strip=True) for td in row.find_all('td')])
                    if headers or data_rows:
                        tables.append({"headers": headers, "rows": data_rows})

        except ImportError:
            # Fall back to basic HTML stripping
            logger.warning("BeautifulSoup not available, using basic HTML stripping")
            import re
            content = re.sub(r'<script[^>]*>.*?</script>', '', html_content, flags=re.DOTALL)
            content = re.sub(r'<style[^>]*>.*?</style>', '', content, flags=re.DOTALL)
            content = re.sub(r'<[^>]+>', ' ', content)
            content = re.sub(r'\s+', ' ', content).strip()
            headings = self._detect_headings(content)
            tables = []

        # Calculate confidence
        confidence = self._calculate_confidence(content, headings, tables)

        logger.info(f"Extracted {len(content)} chars, {len(headings)} headings from URL")

        return ExtractedContent(
            source_type=self.extractor_type,
            source_path=source,
            raw_text=content,
            structural_hints=headings,
            tables=tables,
            metadata={
                "url": source,
                "extraction_method": "web",
            },
            page_count=1,
            errors=errors,
            confidence=confidence,
        )


class ExtractorFactory:
    """Factory for creating extractors based on source type."""

    _extractors: Dict[ExtractorType, type] = {
        ExtractorType.PDF: PdfExtractor,
        ExtractorType.EXCEL: ExcelExtractor,
        ExtractorType.CSV: CsvExtractor,
        ExtractorType.TEXT: TextExtractor,
        ExtractorType.WEB: WebExtractor,
    }

    @classmethod
    def get_extractor(cls, extractor_type: ExtractorType) -> BaseExtractor:
        """Get extractor instance by type."""
        extractor_class = cls._extractors.get(extractor_type)
        if not extractor_class:
            raise ValueError(f"Unknown extractor type: {extractor_type}")
        return extractor_class()

    @classmethod
    def from_file(cls, file_path: str) -> BaseExtractor:
        """Get appropriate extractor for a file path."""
        ext = Path(file_path).suffix.lower()
        mapping = {
            ".pdf": ExtractorType.PDF,
            ".xlsx": ExtractorType.EXCEL,
            ".xls": ExtractorType.EXCEL,
            ".csv": ExtractorType.CSV,
            ".txt": ExtractorType.TEXT,
            ".md": ExtractorType.TEXT,
            ".text": ExtractorType.TEXT,
        }
        extractor_type = mapping.get(ext)
        if not extractor_type:
            raise ValueError(f"Unsupported file extension: {ext}")
        return cls.get_extractor(extractor_type)

    @classmethod
    def from_url(cls, url: str) -> BaseExtractor:
        """Get web extractor for URL."""
        return cls.get_extractor(ExtractorType.WEB)

    @classmethod
    def register_extractor(cls, extractor_type: ExtractorType, extractor_class: type) -> None:
        """Register a custom extractor type."""
        cls._extractors[extractor_type] = extractor_class
